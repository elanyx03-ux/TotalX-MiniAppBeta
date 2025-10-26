import os
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from openpyxl import Workbook, load_workbook

# 🔐 Carica variabili d'ambiente
load_dotenv()
TOKEN = os.getenv("TELEGRAM_TOKEN")

# 📁 File condiviso tra gli admin
ADMIN_FILE = "estratto_conto_admin.xlsx"

# 👑 Admin fisso principale
MAIN_ADMIN = "@Elanyx03"
admins = [MAIN_ADMIN]  # Lista admin, modificabile solo da @Elanyx03

# 🧮 Arrotonda i valori a due cifre decimali
def round_decimal(value):
    return float(Decimal(value).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

# 📄 Crea o carica un file Excel
def load_or_create_file(filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["User", "Movimento", "Data"])
        wb.save(filename)
    return wb, ws

# 💾 Salva un movimento nel file corretto
def salva_movimento(username, valore, admin_mode=False):
    filename = ADMIN_FILE if admin_mode else f"Movimenti_{username}.xlsx"
    wb, ws = load_or_create_file(filename)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([username, round_decimal(valore), now])
    wb.save(filename)

# 📖 Leggi i movimenti dal file
def leggi_movimenti(username, admin_mode=False):
    filename = ADMIN_FILE if admin_mode else f"Movimenti_{username}.xlsx"
    wb, ws = load_or_create_file(filename)
    movimenti = [(row[0], row[1], row[2]) for row in ws.iter_rows(min_row=2, values_only=True)]
    return movimenti

# 💰 Calcola saldo totale
def estratto_conto(username, admin_mode=False):
    movimenti = leggi_movimenti(username, admin_mode)
    totale_entrate = sum(m[1] for m in movimenti if m[1] > 0)
    totale_uscite = sum(m[1] for m in movimenti if m[1] < 0)
    saldo = totale_entrate + totale_uscite
    return movimenti, totale_entrate, totale_uscite, saldo

# ⚙️ COMANDI BOT

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Ciao! Sono TotalX Pro.\n"
        "Ecco i comandi disponibili:\n\n"
        "📥 /add <numero> - aggiunge un’entrata\n"
        "📤 /subtract <numero> - aggiunge un’uscita\n"
        "💰 /total - mostra il saldo\n"
        "📊 /report - mostra tutti i movimenti\n"
        "📂 /export - ricevi il file Excel\n"
        "↩️ /undo - annulla l’ultima operazione\n"
        "🧾 /reset - azzera tutto e crea un nuovo foglio\n\n"
        "👑 Solo admin:\n"
        "/setadmin @username - aggiunge o rimuove un admin\n"
        "/adminlist - mostra gli admin attuali"
    )

async def add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        value = round_decimal(float(context.args[0].replace(",", ".")))
        username = update.message.from_user.username or update.message.from_user.first_name
        admin_mode = f"@{username}" in admins
        salva_movimento(username, value, admin_mode)
        _, _, _, saldo = estratto_conto(username, admin_mode)
        await update.message.reply_text(f"✅ Entrata registrata: +{value}\n💰 Saldo attuale: {saldo}")
    except (IndexError, ValueError):
        await update.message.reply_text("⚠️ Usa il comando così: /add 100 o /add 0,05")

async def subtract(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        value = round_decimal(float(context.args[0].replace(",", ".")))
        username = update.message.from_user.username or update.message.from_user.first_name
        admin_mode = f"@{username}" in admins
        salva_movimento(username, -value, admin_mode)
        _, _, _, saldo = estratto_conto(username, admin_mode)
        await update.message.reply_text(f"✅ Uscita registrata: -{value}\n💰 Saldo attuale: {saldo}")
    except (IndexError, ValueError):
        await update.message.reply_text("⚠️ Usa il comando così: /subtract 50 o /subtract 0,07")

async def total(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    admin_mode = f"@{username}" in admins
    _, _, _, saldo = estratto_conto(username, admin_mode)
    await update.message.reply_text(f"💰 Saldo totale: {saldo}")

async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    admin_mode = f"@{username}" in admins
    movimenti, entrate, uscite, saldo = estratto_conto(username, admin_mode)
    if not movimenti:
        await update.message.reply_text("📭 Nessun movimento registrato.")
        return
    report_text = "📊 Estratto Conto:\n\n"
    for m in movimenti:
        tipo = "Entrata" if m[1] > 0 else "Uscita"
        report_text += f"{tipo}: {m[1]} ({m[0]} - {m[2]})\n"
    report_text += f"\nTotale Entrate: {entrate}\nTotale Uscite: {uscite}\n💰 Saldo Totale: {saldo}"
    await update.message.reply_text(report_text)

async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    admin_mode = f"@{username}" in admins
    filename = ADMIN_FILE if admin_mode else f"Movimenti_{username}.xlsx"
    with open(filename, "rb") as file:
        await update.message.reply_document(file, filename=filename)

async def undo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    admin_mode = f"@{username}" in admins
    filename = ADMIN_FILE if admin_mode else f"Movimenti_{username}.xlsx"
    wb, ws = load_or_create_file(filename)
    if ws.max_row > 1:
        ws.delete_rows(ws.max_row)
        wb.save(filename)
        await update.message.reply_text("↩️ Ultima operazione annullata.")
    else:
        await update.message.reply_text("⚠️ Nessuna operazione da annullare.")

async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    admin_mode = f"@{username}" in admins
    filename = ADMIN_FILE if admin_mode else f"Movimenti_{username}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["User", "Movimento", "Data"])
    wb.save(filename)
    await update.message.reply_text("🧾 Foglio azzerato. Nuovo file creato con saldo 0.")

async def setadmin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    user_tag = f"@{username}"
    if user_tag != MAIN_ADMIN:
        await update.message.reply_text("❌ Solo l’admin principale può modificare gli admin.")
        return
    try:
        target = context.args[0]
        if target == MAIN_ADMIN:
            await update.message.reply_text("⚠️ Non puoi modificare l’admin principale.")
            return
        if target in admins:
            admins.remove(target)
            await update.message.reply_text(f"🗑️ {target} rimosso dagli admin.")
        else:
            admins.append(target)
            await update.message.reply_text(f"✅ {target} aggiunto come admin.")
    except IndexError:
        await update.message.reply_text("⚠️ Usa /setadmin @username")

async def adminlist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.message.from_user.username or update.message.from_user.first_name
    if f"@{username}" not in admins:
        await update.message.reply_text("❌ Solo gli admin possono vedere la lista admin.")
        return
    await update.message.reply_text("👑 Lista admin:\n" + "\n".join(admins))

# 🚀 Avvio bot
def main():
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add", add))
    app.add_handler(CommandHandler("subtract", subtract))
    app.add_handler(CommandHandler("total", total))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("export", export))
    app.add_handler(CommandHandler("undo", undo))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(CommandHandler("setadmin", setadmin))
    app.add_handler(CommandHandler("adminlist", adminlist))

    print("🤖 Bot avviato correttamente...")
    app.run_polling()

if __name__ == "__main__":
    main()
